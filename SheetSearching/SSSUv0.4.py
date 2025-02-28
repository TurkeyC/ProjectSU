import pandas as pd
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk
import sys
import os
import glob
from pathlib import Path

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath("MyBuild")

    return os.path.join(base_path, relative_path)

def search_excel_files(file_paths, search_term, case_sensitive=False):
    try:
        all_results = {}

        for file_path in file_paths:
            try:
                # 获取文件名(不带路径)
                file_name = Path(file_path).name
                file_results = {}
                success = False

                # 定义要尝试的引擎和选项组合
                engines_and_options = [
                    # 首先尝试 openpyxl 标准模式
                    {'engine': 'openpyxl', 'options': {}},
                    # 然后尝试 openpyxl 的只读模式，忽略样式
                    {'engine': 'openpyxl', 'options': {'read_only': True}},
                    # 尝试 xlrd 引擎
                    {'engine': 'xlrd', 'options': {}},
                    # 尝试 pyxlsb 引擎 (适用于 .xlsb 文件)
                    {'engine': 'pyxlsb', 'options': {}},
                    # 最后尝试多种 'odf' 引擎（适用于 OpenOffice 文件）
                    {'engine': 'odf', 'options': {}}
                ]

                for engine_config in engines_and_options:
                    if success:
                        break

                    engine = engine_config['engine']
                    options = engine_config['options']

                    try:
                        # 尝试使用当前引擎获取所有表格名
                        if engine == 'openpyxl' and options.get('read_only', False):
                            # 对于只读模式特别处理
                            from openpyxl import load_workbook
                            wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
                            sheet_names = wb.sheetnames
                        else:
                            # 尝试直接使用 pandas 获取表格名
                            try:
                                xl = pd.ExcelFile(file_path, engine=engine, **options)
                                sheet_names = xl.sheet_names
                            except:
                                # 跳过这个引擎组合
                                continue

                        for sheet_name in sheet_names:
                            try:
                                # 读取表格内容
                                read_options = {
                                    'io': file_path,
                                    'sheet_name': sheet_name,
                                    'engine': engine,
                                    'keep_default_na': False,
                                    'na_values': [],
                                    'header': None
                                }
                                # 合并额外选项
                                read_options.update(options)

                                # 尝试读取并处理异常
                                df = pd.read_excel(**read_options)

                                # 根据大小写敏感设置进行搜索
                                if case_sensitive:
                                    result = df[df.apply(lambda row: row.astype(str).str.contains(search_term, regex=False).any(), axis=1)]
                                else:
                                    result = df[df.apply(lambda row: row.astype(str).str.contains(search_term, case=False, regex=False).any(), axis=1)]

                                if not result.empty:
                                    file_results[sheet_name] = result
                                    success = True
                            except Exception as sheet_error:
                                # 继续尝试其他表格
                                continue

                        # 如果成功读取了所有表格，跳出引擎尝试循环
                        if success:
                            break

                    except Exception:
                        # 尝试下一个引擎组合
                        continue

                # 处理结果
                if file_results:
                    all_results[file_name] = file_results
                else:
                    all_results[file_name] = {"error": "无法读取文件内容或未找到匹配结果"}

            except Exception as e:
                all_results[file_name] = {"error": f"处理文件出错: {str(e)}"}

        return all_results
    except Exception as e:
        return {"error": str(e)}

class ExcelSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel多表搜索工具")
        self.root.geometry("800x600")

        # 文件路径选择
        file_frame = tk.Frame(root)
        file_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(file_frame, text="Excel文件:").pack(side=tk.LEFT)
        self.file_paths = []
        self.files_listbox = tk.Listbox(file_frame, width=50, height=3)
        self.files_listbox.pack(side=tk.LEFT, padx=5, fill=tk.BOTH, expand=True)

        file_buttons_frame = tk.Frame(file_frame)
        file_buttons_frame.pack(side=tk.LEFT)

        tk.Button(file_buttons_frame, text="浏览文件...", command=self.browse_files).pack(fill=tk.X, pady=2)
        tk.Button(file_buttons_frame, text="浏览文件夹...", command=self.browse_folder).pack(fill=tk.X, pady=2)
        tk.Button(file_buttons_frame, text="清除列表", command=self.clear_files).pack(fill=tk.X, pady=2)

        # 搜索框
        search_frame = tk.Frame(root)
        search_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(search_frame, text="搜索内容:").pack(side=tk.LEFT)
        self.search_term = tk.StringVar()
        tk.Entry(search_frame, textvariable=self.search_term, width=50).pack(side=tk.LEFT, padx=5)

        # 选项
        options_frame = tk.Frame(root)
        options_frame.pack(fill=tk.X, padx=20, pady=5)

        self.case_sensitive = tk.BooleanVar()
        tk.Checkbutton(options_frame, text="区分大小写", variable=self.case_sensitive).pack(side=tk.LEFT)

        # 搜索按钮
        tk.Button(options_frame, text="搜索", command=self.search).pack(side=tk.LEFT, padx=10)

        # 结果显示区域
        result_frame = tk.Frame(root)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

        # 结果标签页 - 双层
        self.result_notebook = ttk.Notebook(result_frame)
        self.result_notebook.pack(fill=tk.BOTH, expand=True)

        # 状态栏
        self.status = tk.StringVar()
        tk.Label(root, textvariable=self.status, bd=1, relief=tk.SUNKEN, anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X)

    def browse_files(self):
        filenames = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx *.xls")])
        if filenames:
            for filename in filenames:
                if filename not in self.file_paths:
                    self.file_paths.append(filename)
                    self.files_listbox.insert(tk.END, Path(filename).name)

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            excel_files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
            for file in excel_files:
                if file not in self.file_paths:
                    self.file_paths.append(file)
                    self.files_listbox.insert(tk.END, Path(file).name)

    def clear_files(self):
        self.file_paths = []
        self.files_listbox.delete(0, tk.END)

    def search(self):
        # 清除之前的结果
        for tab in self.result_notebook.winfo_children():
            tab.destroy()

        file_paths = self.file_paths
        search_term = self.search_term.get()

        if not file_paths or not search_term:
            self.status.set("请选择Excel文件并提供搜索内容")
            return

        self.status.set("正在搜索...")
        self.root.update()

        results = search_excel_files(file_paths, search_term, self.case_sensitive.get())

        if "error" in results:
            self.status.set(f"错误: {results['error']}")
            return

        if not results:
            self.status.set("未找到匹配内容")
            return

        # 显示结果
        total_files = len(results)
        total_sheets = 0
        total_rows = 0

        for file_name, file_results in results.items():
            # 为每个文件创建一个标签页
            file_tab = ttk.Frame(self.result_notebook)

            # 跳过错误文件的行计数
            if "error" not in file_results:
                file_sheet_count = len(file_results)
                file_row_count = sum(len(df) for df in file_results.values())
                self.result_notebook.add(file_tab, text=f"{file_name} ({file_sheet_count}表/{file_row_count}行)")

                # 为文件内的每个sheet创建子标签页
                sheet_notebook = ttk.Notebook(file_tab)
                sheet_notebook.pack(fill=tk.BOTH, expand=True)

                for sheet_name, df in file_results.items():
                    sheet_tab = ttk.Frame(sheet_notebook)
                    sheet_notebook.add(sheet_tab, text=f"{sheet_name} ({len(df)})")

                    text_area = scrolledtext.ScrolledText(sheet_tab, wrap=tk.WORD)
                    text_area.pack(fill=tk.BOTH, expand=True)
                    text_area.insert(tk.END, df.to_string())
                    text_area.config(state=tk.DISABLED)

                    total_rows += len(df)

                total_sheets += file_sheet_count
            else:
                self.result_notebook.add(file_tab, text=file_name)
                text_area = scrolledtext.ScrolledText(file_tab, wrap=tk.WORD)
                text_area.pack(fill=tk.BOTH, expand=True)
                text_area.insert(tk.END, f"处理文件时出错: {file_results['error']}")
                text_area.config(state=tk.DISABLED)

        self.status.set(f"在 {total_files} 个文件的 {total_sheets} 个表中找到 {total_rows} 行匹配内容")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSearchApp(root)
    root.mainloop()